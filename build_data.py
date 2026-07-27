#!/usr/bin/env python3
"""
Construit data.json (multi-domaines GDP / SPI_MARO) a partir des exports bruts.

Usage :
    python3 build_data.py Export_GDP.xlsx Export_SPI_MARO.xlsx [--out data.json] [--today AAAA-MM-JJ]

Les deux exports viennent de systemes differents avec des colonnes differentes.
Le mapping colonne -> champ dashboard est explicite ci-dessous. Les lignes
marquees "HYPOTHESE" sont des choix faits par deduction sur les donnees, pas
des infos confirmees par l'utilisateur -- a corriger ici si besoin, tout le
reste du dashboard s'adapte automatiquement.
"""
import sys
import json
import argparse
from datetime import date, datetime
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl requis : pip install openpyxl")

TRANCHES = ['0-28 jours', '29-60 jours', '61-90 jours', '91-120 jours', '120+ jours']


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.startswith('0001-01-01'):
        return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


def iso(d):
    return d.isoformat() if d else None


def bucket_tranche(jours):
    if jours <= 28:
        return TRANCHES[0]
    if jours <= 60:
        return TRANCHES[1]
    if jours <= 90:
        return TRANCHES[2]
    if jours <= 120:
        return TRANCHES[3]
    return TRANCHES[4]


def derive_dates(date_ouverture, date_echeance, statut, date_cloture, today):
    """Reproduit la logique jours_ecoules / tranche / en_retard / jours_retard
    utilisee par le dashboard (meme convention que le jeu de demo)."""
    if statut == 'Clôturé' and date_cloture and date_ouverture:
        jours_ecoules = (date_cloture - date_ouverture).days
        ref = date_cloture
    elif date_ouverture:
        jours_ecoules = (today - date_ouverture).days
        ref = today
    else:
        jours_ecoules = 0
        ref = today
    jours_ecoules = max(jours_ecoules, 0)
    if date_echeance and ref:
        en_retard = 'Oui' if ref > date_echeance else 'Non'
        jours_retard = max((ref - date_echeance).days, 0)
    else:
        en_retard = 'Non'
        jours_retard = 0
    return jours_ecoules, bucket_tranche(jours_ecoules), en_retard, jours_retard


def read_rows(path, sheet='Données'):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return idx, rows


def build_meta(dossiers, perimetres_field_label):
    zones = sorted({d['zone'] for d in dossiers if d.get('zone')})
    activites = sorted({d['activite'] for d in dossiers if d.get('activite')})
    perimetres = [{'key': 'all', 'label': 'Toutes les RPL', 'activite': None}]
    perimetres += [{'key': a, 'label': a, 'activite': a} for a in activites]
    return {
        'perimetres': perimetres,
        'zones': zones,
        'perimetre_label': perimetres_field_label,
    }


WEEKS_WINDOW = 104  # fenetre du graphique de tendance : ~2 ans glissants avant `today`.
# Les dossiers plus anciens que ca restent comptes normalement partout ailleurs
# (KPI, tranches, retard...) ; seule la courbe hebdomadaire est bornee, sinon
# quelques dossiers tres anciens (2018-2021, une quinzaine de cas) etirent
# l'axe des semaines sur 400+ points pour rien.
def build_weeks(dossiers, today):
    dates = [d['date_ouverture'] for d in dossiers if d.get('date_ouverture')]
    if not dates:
        return []
    earliest = min(datetime.strptime(d, '%Y-%m-%d').date() for d in dates)
    start = today - __import__('datetime').timedelta(weeks=WEEKS_WINDOW)
    start = max(start, earliest)
    # aligne sur le lundi
    start -= __import__('datetime').timedelta(days=start.weekday())
    weeks = []
    cur = start
    while cur <= today:
        iso_cal = cur.isocalendar()
        weeks.append({'start': cur.isoformat(), 'label': f'S{iso_cal[1]:02d}-{iso_cal[0]}'})
        cur += __import__('datetime').timedelta(days=7)
    return weeks


# =========================================================================
# GDP
# =========================================================================
def build_gdp(path, today):
    idx, rows = read_rows(path)

    def val(r, col):
        return r[idx[col]]

    dossiers = []
    for r in rows:
        oeie = val(r, 'OEIE')
        if not oeie:
            continue

        date_ouverture = to_date(val(r, 'DTC'))          # HYPOTHESE : DTC = date de creation/ouverture (100% renseigne)
        date_echeance = to_date(val(r, 'DLR'))            # HYPOTHESE : DLR = delai limite de realisation (100% renseigne)

        etat = val(r, 'ETAT')
        # HYPOTHESE : ETAT in {3,4} correle a 97-100% avec une date de fin
        # (DATEETAT3 / DateEtat4) renseignee -> traite comme Cloture.
        # ETAT in {1,2,5} n'a jamais ces dates -> traite comme Ouvert.
        # A CONFIRMER avec l'utilisateur.
        statut = 'Clôturé' if etat in (3, 4) else 'Ouvert'
        date_cloture = to_date(val(r, 'DATEETAT3')) or to_date(val(r, 'DateEtat4')) if statut == 'Clôturé' else None

        client = (val(r, 'titulaire') or '').strip() or None   # HYPOTHESE : titulaire = client
        responsable = (val(r, 'Equipe') or '').strip() or None
        activite = (val(r, 'codeservice') or '').strip() or None  # HYPOTHESE : codeservice = RPL/activite
        if activite == '.':  # artefact d'export (12 lignes) -> traite comme non renseigne
            activite = None
        codecomune = val(r, 'codecomune')
        zone = str(codecomune)[:2] if codecomune else None

        demande = (val(r, 'demande') or '').strip()
        num_commande = demande or None
        sans_commande = 'Oui' if not num_commande else 'Non'
        sans_responsable = 'Oui' if not responsable else 'Non'

        jours_ecoules, tranche, en_retard, jours_retard = derive_dates(
            date_ouverture, date_echeance, statut, date_cloture, today)

        dossiers.append({
            'dossier': str(oeie),
            'client': client,
            'activite': activite,
            'zone': zone,
            'responsable': responsable,
            'date_ouverture': iso(date_ouverture),
            'date_echeance': iso(date_echeance),
            'statut': statut,
            'date_cloture': iso(date_cloture),
            'jours_ecoules': jours_ecoules,
            'tranche': tranche,
            'en_retard': en_retard,
            'jours_retard': jours_retard,
            'num_commande': num_commande,
            'sans_commande': sans_commande,
            'sans_responsable': sans_responsable,
            'type_etat': (val(r, 'typePoi') or 'Non renseigné'),  # pour le graphique "par etat"
            'ville': val(r, 'ville'),
        })

    meta = build_meta(dossiers, 'RPL (codeservice)')
    weeks = build_weeks(dossiers, today)
    return {'meta': meta, 'dossiers': dossiers, 'weeks': weeks}


# =========================================================================
# SPI_MARO
# =========================================================================
# Regroupement des Type_activite en etapes du "suivi" (etude -> commande de
# materiel -> travaux [par sous-type] -> controle -> cloture), tel que decrit
# par l'utilisateur. Les "?" de la proposition sont places ici avec le
# meilleur jugement possible -- a corriger si besoin.
ACTIVITY_STAGE = {
    # --- Etude ---
    'A/R Etude fibre': 'Étude',
    'Valorisation Etude Fibre liste de MA&MO': 'Étude',
    'Valo Etude Fibre liste de MA&MO (ETR)': 'Étude',
    'Réalisation étude Fibre': 'Étude',
    'Réalisation étude Fibre (ETR)': 'Étude',
    'Réaliser Etude': 'Étude',
    'Etude Forfait RR': 'Étude',
    'Saisie RDT étude Fibre': 'Étude',
    'Identification CAFF référent': 'Étude',
    'Validation RDT étude Fibre': 'Étude',
    'Valider Etude': 'Étude',
    'Identification ETR': 'Étude',
    'Identifier les acteurs': 'Étude',
    'Confirmer prise en charge': 'Étude',
    'Identification Pilote': 'Étude',
    'Identification CAFF': 'Étude',
    'A/R Forfait RR': 'Étude',
    'Ajout des options': 'Étude',
    # --- Commande de matériel ---
    'Commande de matériels RR Fibre (ETR)': 'Commande de matériel',
    'Commande de matériels RR fibre': 'Commande de matériel',
    'Saisie matériels RR Fibre (ETR)': 'Commande de matériel',
    # --- Travaux (sous-types) ---
    'A/R Travaux Fibre': 'Travaux — Fibre',
    'Travaux Fibre': 'Travaux — Fibre',
    'Saisie RDT Tvx Fibre': 'Travaux — Fibre',
    'Planifier travaux': 'Travaux — Planification',
    'Planification des travaux Fibre': 'Travaux — Planification',
    'Réaliser travaux BL': 'Travaux — BL',
    'Réaliser travaux GC': 'Travaux — GC',
    'Travaux Forfait RR': 'Travaux — Forfait RR',
    'Saisie acompte RDT Forfait RR': 'Travaux — Forfait RR',
    # --- Contrôle ---
    'Contrôle travaux, SI et DOE Fibre (ETR)': 'Contrôle',
    "Contrôler l'opération": 'Contrôle',
    'Validation RDT Tvx Fibre': 'Contrôle',
    'Validation acompte RDT Forfait RR': 'Contrôle',
    'Validation solde RDT Forfait RR': 'Contrôle',
    # --- Clôture ---
    "Clôturer l'opération": 'Clôture',
    'Clôture du projet': 'Clôture',
    'Fin du dossier Forfait RR': 'Clôture',
    'Saisie solde RDT Forfait RR': 'Clôture',
}


def build_spi_maro(path, today):
    idx, rows = read_rows(path)

    def val(r, col):
        return r[idx[col]]

    # dedoublonnage par Code_Operation_global : on garde la ligne avec la
    # Date_de_lancement la plus recente (demande utilisateur).
    groups = defaultdict(list)
    for r in rows:
        cog = val(r, 'Code_Operation_global')
        if not cog:
            continue
        groups[cog].append(r)

    unmapped_activities = Counter()
    dossiers = []
    for cog, group_rows in groups.items():
        group_rows.sort(key=lambda r: to_date(val(r, 'Date_de_lancement')) or date.min, reverse=True)
        r = group_rows[0]

        type_activite = (val(r, 'Type_activite') or '').strip()
        stage = ACTIVITY_STAGE.get(type_activite)
        if stage is None and type_activite:
            unmapped_activities[type_activite] += 1
            stage = 'Non classé'

        date_ouverture = to_date(val(r, 'Date_de_depot_du_projet'))  # HYPOTHESE : depot = ouverture
        date_echeance = to_date(val(r, 'DLR'))                       # meme convention que GDP

        statut = 'Clôturé' if stage == 'Clôture' else 'Ouvert'       # derive du stade d'activite
        date_cloture = to_date(val(r, 'Date_de_fin_reelle')) if statut == 'Clôturé' else None

        code_insee = val(r, 'Code_INSEE__projet')
        zone = str(code_insee)[:2] if code_insee else None
        activite = (val(r, 'rpl') or '').strip() or None
        responsable = (val(r, 'Poste_de_travail') or '').strip() or None  # HYPOTHESE : poste de travail = responsable
        client = None  # AUCUNE colonne candidate identifiee -- a fournir

        num_commande = (val(r, 'Identifiant_du_projet') or '').strip() or None
        sans_commande = 'Oui' if not num_commande else 'Non'
        sans_responsable = 'Oui' if not responsable else 'Non'

        jours_ecoules, tranche, en_retard, jours_retard = derive_dates(
            date_ouverture, date_echeance, statut, date_cloture, today)

        dossiers.append({
            'dossier': str(cog),
            'client': client,
            'activite': activite,
            'zone': zone,
            'responsable': responsable,
            'date_ouverture': iso(date_ouverture),
            'date_echeance': iso(date_echeance),
            'statut': statut,
            'date_cloture': iso(date_cloture),
            'jours_ecoules': jours_ecoules,
            'tranche': tranche,
            'en_retard': en_retard,
            'jours_retard': jours_retard,
            'num_commande': num_commande,
            'sans_commande': sans_commande,
            'sans_responsable': sans_responsable,
            'type_etat': stage or 'Non renseigné',  # pour le graphique "par etat" (etape du suivi)
            'type_si': val(r, 'TypeSI'),
        })

    if unmapped_activities:
        print('!! Activites SPI_MARO non classees (traitees en "Non classé") :', file=sys.stderr)
        for a, c in unmapped_activities.most_common():
            print(f'   {c:4d}  {a}', file=sys.stderr)

    meta = build_meta(dossiers, 'RPL')
    weeks = build_weeks(dossiers, today)
    return {'meta': meta, 'dossiers': dossiers, 'weeks': weeks}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('gdp_path')
    ap.add_argument('spi_maro_path')
    ap.add_argument('--out', default='data.json')
    ap.add_argument('--today')
    args = ap.parse_args()

    today = datetime.strptime(args.today, '%Y-%m-%d').date() if args.today else date.today()

    gdp = build_gdp(args.gdp_path, today)
    spi_maro = build_spi_maro(args.spi_maro_path, today)

    out = {
        'meta': {
            'generated': today.isoformat(),
            'today': today.isoformat(),
            'objectif_respect': 0.9,
            'marge_tolerance_respect': 0.1,
            'seuil_bas_respect': 0.8,
            'objectif_hors_delai': 0.1,
            'seuil_alerte_hors_delai': 0.2,
        },
        'domains': {
            'GDP': gdp,
            'SPI_MARO': spi_maro,
        },
    }
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=None, separators=(',', ':'))

    print(f'OK -> {args.out}')
    print(f"  GDP      : {len(gdp['dossiers'])} dossiers")
    print(f"  SPI_MARO : {len(spi_maro['dossiers'])} dossiers (dedoublonnes)")


if __name__ == '__main__':
    main()
